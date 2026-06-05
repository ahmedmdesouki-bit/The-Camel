from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT='Arial'
BLUE=Font(name=FONT,color='0000FF')      # inputs
BLACK=Font(name=FONT,color='000000')     # formulas
GREEN=Font(name=FONT,color='008000')     # cross-sheet
HDR=Font(name=FONT,bold=True,color='FFFFFF',size=11)
TITLE=Font(name=FONT,bold=True,size=15,color='1F3864')
SUB=Font(name=FONT,italic=True,color='595959',size=9)
BOLD=Font(name=FONT,bold=True)
HDRFILL=PatternFill('solid',start_color='1F3864')
YEL=PatternFill('solid',start_color='FFF2CC')
GREYFILL=PatternFill('solid',start_color='F2F2F2')
REDFILL=PatternFill('solid',start_color='FCE4E4')
GRNFILL=PatternFill('solid',start_color='E2EFDA')
C=Alignment(horizontal='center',vertical='center',wrap_text=True)
L=Alignment(horizontal='left',vertical='center',wrap_text=True)
thin=Side(style='thin',color='BFBFBF')
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook()

def hdrrow(ws,row,headers,start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=start+i,value=h)
        c.font=HDR;c.fill=HDRFILL;c.alignment=C;c.border=BORD

def setw(ws,widths):
    for col,w in widths.items():
        ws.column_dimensions[col].width=w

# ---------- CONTROL ----------
ws=wb.active; ws.title='Control'
ws['A1']='StockSense v11 — Control Panel'; ws['A1'].font=TITLE
ws['A2']='Yellow cells = update these. Peg is fixed; EGP floats; live prices on Holdings sheet.'; ws['A2'].font=SUB
rows=[
 ('Setting','Value','Notes'),
 ('USD / SAR (peg)',3.75,'Fixed ~3.75 — SAR pegged to USD'),
 ('USD / EGP (float)',49.0,'UPDATE — EGP floats; check live rate'),
 ('Monthly deposit (USD)',100,'Your DCA contribution'),
 ('Cash on hand (USD)',27.82,'Idle cash in Sahm wallet'),
 ('Cash-drag flag at this scale (USD)',100,'Flag if idle cash > 1 monthly deposit (<$1k book)'),
 ('Sector concentration limit',0.40,'🚨 if any sector > 40%'),
]
r=4
for i,(a,b,c) in enumerate(rows):
    ra=ws.cell(row=r+i,column=1,value=a)
    rb=ws.cell(row=r+i,column=2,value=b)
    rc=ws.cell(row=r+i,column=3,value=c)
    if i==0:
        for cc in (ra,rb,rc): cc.font=HDR;cc.fill=HDRFILL;cc.alignment=C;cc.border=BORD
    else:
        ra.font=BOLD;ra.border=BORD;ra.alignment=L
        rb.font=BLUE;rb.border=BORD;rb.alignment=C;rb.fill=YEL
        rc.font=SUB;rc.border=BORD;rc.alignment=L
ws['B7'].number_format='#,##0'; ws['B8'].number_format='$#,##0.00'
ws['B9'].number_format='$#,##0'; ws['B10'].number_format='0%'
ws['B5'].number_format='0.000'; ws['B6'].number_format='0.00'
setw(ws,{'A':34,'B':14,'C':52})
# named refs by cell: SAR=B5,EGP=B6,DEPOSIT=B7,CASH=B8,DRAGFLAG=B9,SECLIMIT=B10

# ---------- HOLDINGS ----------
ws=wb.create_sheet('Holdings')
ws['A1']='Holdings — Live P&L (hard-currency)'; ws['A1'].font=TITLE
ws['A2']='Enter LIVE PRICE (yellow). Everything else calculates. USD book has ~no FX risk for you (SAR pegged).'; ws['A2'].font=SUB
H=['Market','Ticker','Name','Sharia','Sector','Curr','Shares','Buy Px','Cost (local)','Live Px','Mkt Val (local)','FX→USD','Cost USD','Mkt Val USD','P&L USD','P&L %','Weight']
hr=4; hdrrow(ws,hr,H)
data=[
 ('US','SCHD','Schwab US Dividend Equity ETF','❌ No','Diversified ETF','USD',3,32.28),
 ('US','SCHX','Schwab US Large-Cap ETF','❌ No','Diversified ETF','USD',1,29.76),
]
first=hr+1
for i,(mk,tk,nm,sh,sec,cur,shares,buy) in enumerate(data):
    r=first+i
    ws.cell(r,1,mk).font=BLACK
    ws.cell(r,2,tk).font=BLUE
    ws.cell(r,3,nm).font=BLUE
    cs=ws.cell(r,4,sh); cs.font=BLACK; cs.fill=REDFILL if '❌' in sh else GRNFILL
    ws.cell(r,5,sec).font=BLUE
    ws.cell(r,6,cur).font=BLUE
    ws.cell(r,7,shares).font=BLUE
    ws.cell(r,8,buy).font=BLUE; ws.cell(r,8).number_format='$#,##0.00'
    ws.cell(r,9,f'=G{r}*H{r}').font=BLACK; ws.cell(r,9).number_format='$#,##0.00'
    lv=ws.cell(r,10,f'=H{r}'); lv.font=BLUE; lv.fill=YEL; lv.number_format='$#,##0.00'  # default to buy px
    ws.cell(r,11,f'=G{r}*J{r}').font=BLACK; ws.cell(r,11).number_format='$#,##0.00'
    ws.cell(r,12,f'=IF(F{r}="USD",1,IF(F{r}="SAR",1/Control!$B$5,IF(F{r}="EGP",1/Control!$B$6,1)))').font=GREEN; ws.cell(r,12).number_format='0.0000'
    ws.cell(r,13,f'=I{r}*L{r}').font=BLACK; ws.cell(r,13).number_format='$#,##0.00'
    ws.cell(r,14,f'=K{r}*L{r}').font=BLACK; ws.cell(r,14).number_format='$#,##0.00'
    ws.cell(r,15,f'=N{r}-M{r}').font=BLACK; ws.cell(r,15).number_format='$#,##0.00;($#,##0.00);-'
    ws.cell(r,16,f'=IF(M{r}=0,0,O15)').font=BLACK  # placeholder, fix below
    ws.cell(r,16,f'=IF(M{r}=0,0,(N{r}-M{r})/M{r})').font=BLACK; ws.cell(r,16).number_format='0.0%'
    ws.cell(r,17,f'=IF($N${first+len(data)}=0,0,N{r}/$N${first+len(data)+1})').font=BLACK; ws.cell(r,17).number_format='0.0%'
    for cc in range(1,18): ws.cell(r,cc).border=BORD; ws.cell(r,cc).alignment=C
last=first+len(data)-1
tr=last+1
ws.cell(tr,3,'TOTAL').font=BOLD
for col,lett in [(9,'I'),(11,'K'),(13,'M'),(14,'N'),(15,'O')]:
    pass
ws.cell(tr,13,f'=SUM(M{first}:M{last})').font=BOLD; ws.cell(tr,13).number_format='$#,##0.00'
ws.cell(tr,14,f'=SUM(N{first}:N{last})').font=BOLD; ws.cell(tr,14).number_format='$#,##0.00'
ws.cell(tr,15,f'=SUM(O{first}:O{last})').font=BOLD; ws.cell(tr,15).number_format='$#,##0.00;($#,##0.00);-'
ws.cell(tr,16,f'=IF(M{tr}=0,0,O{tr}/M{tr})').font=BOLD; ws.cell(tr,16).number_format='0.0%'
ws.cell(tr,17,f'=SUM(Q{first}:Q{last})').font=BOLD; ws.cell(tr,17).number_format='0.0%'
for cc in range(1,18):
    ws.cell(tr,cc).fill=GREYFILL; ws.cell(tr,cc).border=BORD; ws.cell(tr,cc).alignment=C
# summary block
sr=tr+2
ws.cell(sr,3,'Portfolio market value (USD)').font=BOLD
ws.cell(sr,5,f'=M{tr}').font=GREEN; ws.cell(sr,5).number_format='$#,##0.00'
ws.cell(sr+1,3,'Cash on hand (USD)').font=BOLD
ws.cell(sr+1,5,'=Control!B8').font=GREEN; ws.cell(sr+1,5).number_format='$#,##0.00'
ws.cell(sr+2,3,'Total book value (USD)').font=BOLD
ws.cell(sr+2,5,f'=E{sr}+E{sr+1}').font=BLACK; ws.cell(sr+2,5).number_format='$#,##0.00'
ws.cell(sr+3,3,'Cash drag %').font=BOLD
ws.cell(sr+3,5,f'=IF(E{sr+2}=0,0,E{sr+1}/E{sr+2})').font=BLACK; ws.cell(sr+3,5).number_format='0.0%'
ws.cell(sr+4,3,'Cash-drag flag').font=BOLD
ws.cell(sr+4,5,f'=IF(Control!B8>Control!B9,"🚨 Idle cash too high — deploy","✅ OK")').font=BLACK
setw(ws,{'A':7,'B':9,'C':30,'D':9,'E':16,'F':7,'G':9,'H':10,'I':13,'J':10,'K':14,'L':9,'M':12,'N':13,'O':12,'P':9,'Q':9})

# ---------- WATCHLIST ----------
ws=wb.create_sheet('Watchlist')
ws['A1']='Watchlist — compliant ETFs + single names'; ws['A1'].font=TITLE
ws['A2']='Affordability auto-calcs from your cash + deposit (Control). Run "Analyze [ticker]" before any buy.'; ws['A2'].font=SUB
H=['Ticker','Name','Type','Sharia','~Price USD','Buy now?','Buy w/ 1 deposit?','Moat 1-5','Status','Notes']
hr=4; hdrrow(ws,hr,H)
wl=[
 ('SPUS','SP Funds S&P 500 Sharia','Compliant ETF','✅ Yes',59.00,'','','','Staged','Cleanest 1:1 swap; liquid; monthly dividend'),
 ('HLAL','Wahed FTSE USA Shariah','Compliant ETF','✅ Yes',71.11,'','','','Staged','Broader large+mid; ~$892M AUM'),
 ('MNZL','Manzil Russell Halal USA','Compliant ETF','✅ Yes',60.29,'','','','Staged','Cheapest fee 0.40%; new/thin — wider spreads'),
 ('AAPL','Apple','Single name','⏳','','','','','Pending','Confirmed SCHX exposure'),
 ('DELL','Dell','Single name','⏳','','','','','Pending','Partial exposure'),
 ('FTNT','Fortinet','Single name','⏳','','','','','Pending','Partial exposure'),
 ('AMD','AMD','Single name','⏳','','','','','Pending','Confirmed exposure'),
 ('TXN','Texas Instruments','Single name','⏳','','','','','Pending','Partial exposure'),
 ('JBL','Jabil','Single name','⏳','','','','','Pending','Partial exposure'),
 ('ORCL','Oracle','Single name','⏳','','','','','Pending','Confirmed exposure'),
 ('DDOG','Datadog','Single name','⏳','','','','','Pending','Confirmed exposure'),
]
fr=hr+1
for i,row in enumerate(wl):
    r=fr+i
    for j,val in enumerate(row):
        c=ws.cell(r,j+1,val if val!='' else None)
        c.border=BORD;c.alignment=C if j!=9 else L
        if j in (0,1,2,9): c.font=BLACK if j in(0,1,2) else SUB
        if j==1: c.alignment=L
        if j==3:
            c.font=BLACK
            if '✅' in str(val): c.fill=GRNFILL
            elif '⏳' in str(val): c.fill=YEL
        if j==4 and val!='':
            c.font=BLUE; c.fill=YEL; c.number_format='$#,##0.00'
        if j in (7,): c.font=BLUE
    # affordability formulas only where price present
    if row[4]!='':
        ws.cell(r,6,f'=IF(E{r}="","",INT(Control!$B$8/E{r}))').font=GREEN
        ws.cell(r,7,f'=IF(E{r}="","",INT((Control!$B$8+Control!$B$7)/E{r}))').font=GREEN
setw(ws,{'A':9,'B':22,'C':14,'D':9,'E':12,'F':10,'G':16,'H':9,'I':10,'J':40})

# ---------- SHARIA RE-SCREEN LOG ----------
ws=wb.create_sheet('Sharia ReScreen')
ws['A1']='Sharia Re-Screen Log — run quarterly'; ws['A1'].font=TITLE
ws['A2']='Hard wall. Business screen first, then ratios. Thresholds: debt/mktcap <33%, cash+sec/mktcap <33%, non-compliant income <5%.'; ws['A2'].font=SUB
H=['Date','Ticker','Business screen','Debt/MktCap','Cash+Sec/MktCap','Non-compl. income','Verdict','Source','Next review']
hr=4; hdrrow(ws,hr,H)
seed=[
 ('2026-06-05','SCHD','Conventional index','>33%','—','—','❌ Not compliant','Musaffa May 2026','2026-09'),
 ('2026-06-05','SCHX','Holds financials','>33%','—','—','❌ Not compliant','Musaffa','2026-09'),
 ('2026-06-05','SPUS','Sharia-screened','<33%','<33%','<5%','✅ Compliant','S&P Sharia index','2026-09'),
 ('2026-06-05','HLAL','Sharia-screened','<33%','<33%','<5%','✅ Compliant','FTSE USA Shariah','2026-09'),
 ('2026-06-05','MNZL','Sharia-screened','<33%','<33%','<5%','✅ Compliant','Russell IdealRatings','2026-09'),
]
fr=hr+1
for i,row in enumerate(seed):
    r=fr+i
    for j,val in enumerate(row):
        c=ws.cell(r,j+1,val); c.border=BORD; c.alignment=C; c.font=BLACK
        if j==6:
            c.fill=REDFILL if '❌' in val else GRNFILL
# blank rows for future
for k in range(8):
    r=fr+len(seed)+k
    for j in range(9):
        ws.cell(r,j+1).border=BORD
setw(ws,{'A':12,'B':9,'C':18,'D':12,'E':16,'F':16,'G':17,'H':18,'I':12})

# ---------- RESEARCH LOG ----------
ws=wb.create_sheet('Research Log')
ws['A1']='Research Log — the system\'s memory (one entry per session)'; ws['A1'].font=TITLE
ws['A2']='No position without a written invalidation trigger. Fill Outcome later — that is how the system learns.'; ws['A2'].font=SUB
H=['Date','Market','Ticker','Action','Thesis (1 line)','Invalidation trigger','Profit-take rule','Time stop','Base rate used','Source','Outcome']
hr=4; hdrrow(ws,hr,H)
seed=[
 ('2026-06-05','US','SCHD/SCHX','SHARIA SCAN','Held as core','Failed Sharia gate → plan exit','n/a','Resolve this quarter','n/a','Musaffa','Pivot to SPUS pending'),
]
fr=hr+1
for i,row in enumerate(seed):
    r=fr+i
    for j,val in enumerate(row):
        c=ws.cell(r,j+1,val);c.border=BORD;c.alignment=L;c.font=BLACK
for k in range(14):
    r=fr+len(seed)+k
    for j in range(11): ws.cell(r,j+1).border=BORD
setw(ws,{'A':12,'B':8,'C':10,'D':12,'E':26,'F':26,'G':16,'H':12,'I':16,'J':14,'K':22})

# ---------- DCA / DEPOSITS ----------
ws=wb.create_sheet('DCA Log')
ws['A1']='DCA / Deposit Log — emotion-free deployment'; ws['A1'].font=TITLE
ws['A2']='Log every deposit and buy. Cumulative tracks total contributed.'; ws['A2'].font=SUB
H=['Date','Type','Ticker','Amount USD','Shares','Price','Cumulative deposited','Notes']
hr=4; hdrrow(ws,hr,H)
seed=[
 ('2026-06-01','Buy','SCHD',96.84,3,32.28,'','Day 1'),
 ('2026-06-01','Buy','SCHX',29.76,1,29.76,'','Day 1'),
]
fr=hr+1
for i,row in enumerate(seed):
    r=fr+i
    for j,val in enumerate(row):
        c=ws.cell(r,j+1,val if val!='' else None);c.border=BORD;c.alignment=C;c.font=BLUE if j in(0,1,2,3,4,5,7) else BLACK
        if j==3: c.number_format='$#,##0.00'
        if j==5: c.number_format='$#,##0.00'
    # cumulative deposited (only count Deposit type)
    ws.cell(r,7,f'=SUMIF($B$5:$B{r},"Deposit",$D$5:$D{r})').font=BLACK
    ws.cell(r,7).number_format='$#,##0.00'
for k in range(14):
    r=fr+len(seed)+k
    for j in range(8): ws.cell(r,j+1).border=BORD
    ws.cell(r,7,f'=SUMIF($B$5:$B{r},"Deposit",$D$5:$D{r})').font=BLACK
    ws.cell(r,7).number_format='$#,##0.00'
setw(ws,{'A':12,'B':10,'C':9,'D':12,'E':8,'F':10,'G':18,'H':30})

# ---------- SECTOR RADAR ----------
ws=wb.create_sheet('Sector Radar')
ws['A1']='Sector Radar — concentration guardrail'; ws['A1'].font=TITLE
ws['A2']='Pulls market value by sector from Holdings. 🚨 if any sector exceeds the limit in Control (40%).'; ws['A2'].font=SUB
H=['Sector','Mkt Val USD','Weight','Flag']
hr=4; hdrrow(ws,hr,H)
sectors=['Diversified ETF','Tech','Energy','Consumer Staples','Healthcare','Financials','Other']
fr=hr+1
for i,sec in enumerate(sectors):
    r=fr+i
    ws.cell(r,1,sec).font=BLACK
    ws.cell(r,2,f'=SUMIF(Holdings!$E:$E,A{r},Holdings!$N:$N)').font=GREEN; ws.cell(r,2).number_format='$#,##0.00'
    ws.cell(r,3,f'=IF($B${fr+len(sectors)}=0,0,B{r}/$B${fr+len(sectors)})').font=BLACK; ws.cell(r,3).number_format='0.0%'
    ws.cell(r,4,f'=IF(C{r}>Control!$B$10,"🚨 Over limit","")').font=BLACK
    for cc in range(1,5): ws.cell(r,cc).border=BORD; ws.cell(r,cc).alignment=C
tr=fr+len(sectors)
ws.cell(tr,1,'TOTAL').font=BOLD
ws.cell(tr,2,f'=SUM(B{fr}:B{tr-1})').font=BOLD; ws.cell(tr,2).number_format='$#,##0.00'
ws.cell(tr,3,f'=IF(B{tr}=0,0,SUM(C{fr}:C{tr-1}))').font=BOLD; ws.cell(tr,3).number_format='0.0%'
for cc in range(1,5): ws.cell(tr,cc).fill=GREYFILL; ws.cell(tr,cc).border=BORD; ws.cell(tr,cc).alignment=C
setw(ws,{'A':20,'B':16,'C':10,'D':16})

# ---------- BASE RATE WORKSHEET ----------
ws=wb.create_sheet('Base-Rate Worksheet')
ws['A1']='Base-Rate Worksheet — fill before acting on any signal'; ws['A1'].font=TITLE
ws['A2']='Probability tilt + size, never a buy/sell call. Small sample = low trust.'; ws['A2'].font=SUB
fields=[
 ('Signal','What triggered this?'),
 ('Comparables (N)','How many prior cases? (small N = low trust)'),
 ('Horizon','Window measured (e.g. 3–6 months)'),
 ('Hit rate','% of cases that moved as expected'),
 ('Magnitude','Median + range of the move'),
 ('Already priced in','Honest discount on the obvious read'),
 ('Counter-signals','What would flip this?'),
 ('Overfitting check','Real mechanism or curve fit?'),
 ('→ Probability tilt','Lean, not certainty'),
 ('→ Position size','Shares / $ sized to invalidation distance'),
]
r=4
for i,(a,b) in enumerate(fields):
    ws.cell(r+i,1,a).font=BOLD; ws.cell(r+i,1).fill=GREYFILL; ws.cell(r+i,1).border=BORD; ws.cell(r+i,1).alignment=L
    c=ws.cell(r+i,2,b); c.font=SUB; c.border=BORD; c.alignment=L; c.fill=YEL
setw(ws,{'A':22,'B':70})

wb.save('/sessions/amazing-wizardly-lovelace/mnt/outputs/StockSense_v11_Tracker.xlsx')
print('saved')
